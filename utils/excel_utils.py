import openpyxl
import os

def read_test_data(file_path, sheet_name="Sheet1"):
    """
    Read test data from Excel file
    Returns list of dictionaries where each dictionary represents a test case
    """
    # Check if file exists
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"Excel file not found: {file_path}")
    
    try:
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook[sheet_name]
        
        # Get headers from first row
        headers = [cell.value for cell in sheet[1]]
        
        test_data = []
        for row in range(2, sheet.max_row + 1):
            row_data = {}
            for col, header in enumerate(headers, 1):
                row_data[header] = sheet.cell(row=row, column=col).value
            test_data.append(row_data)
        
        workbook.close()
        return test_data
        
    except Exception as e:
        raise Exception(f"Error reading Excel file: {e}")